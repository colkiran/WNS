
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Charts")

wb.active = wb['Charts']

ws = wb.active

data = [
    ('Products', 'Sales'),
    ('Pepsi', 450),
    ('Coke', 320),
    ('Sprite', 250),
    ('Mirinda', 300),
    ('Fanta', 230),
    ('Thums up', 400),
    ('Slice', 285)
]

print(F"data :{data}")

for row in data:
    ws.append(row)

dataRef = Reference(ws, min_row=2, min_col=2, max_row=8)
labelRef = Reference(ws, min_row=2, max_row=8, min_col=1)

chart =  BarChart()
chart.add_data(dataRef)
chart.set_categories(labelRef)
chart.x_axis.title = "Products"
chart.y_axis.title = "Sales in lakhs"
chart.title = "Baverage Sales"


ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")




