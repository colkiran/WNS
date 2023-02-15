
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()
ws = wb.active              # convert the current worksheet as activesheet
ws.title = "MySheet"        # change the title of the sheet

ws['B5'] = "Hello World"
ws['B6'] = 1234567
ws['B7'] = datetime.now()

wb.save("FirstExcel.xlsx")
